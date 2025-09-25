
###################################################
# settings
###################################################

# the file containing the weapons each operator has access to
operators_file_name = "operators"

# the file containing the attachment overview
attachment_overview_file_name = "attachment_overview"

# the file name of the xlsx output file
xlsx_output_file_name = "rainbow-six-siege-weapon-statistics"

# the directory containing the weapon damage files
weapon_data_dir = "weapons"

# the distance the weapon damage stats start at (usually 0)
first_distance = 0
# the distance the weapon damage stats end at (usually 40 because the of the Shooting Range limit)
last_distance = 40 

# weapon type background colors
weapon_colors = {"AR":"#5083EA",
                 "SMG":"#B6668E",
                 "MP":"#76A5AE",
                 "LMG":"#8771BD",
                 "DMR":"#7CB563",
                 "SR":"#DE2A00",
                 "SG":"#FFBC01",
                 "Slug SG":"#A64E06",
                 "Handgun":"#A3A3A3",
                 "Revolver":"#F48020",
                 "Hand Canon":"#948A54"}

weapon_colors_ = {"AR":"#1f77b4",
                 "SMG":"#e377c2",
                 "MP":"#1abc9c",
                 "LMG":"#9467bd",
                 "DMR":"#2ca02c",
                 "SR":"#d62728",
                 "SG":"#17becf",
                 "Slug SG":"#8c564b",
                 "Handgun":"#7f7f7f",
                 "Revolver":"#ff7f0e",
                 "Hand Canon":"#bcbd22"}

###################################################
# settings end
# don't edit from here on
###################################################

#imports
import os, json, typing, copy, sys, itertools, colorama, sys, colorsys, pandas as pd, numpy as np, io, marshmallow.exceptions
import openpyxl, dataclasses_json, warnings, googleapiclient.http, openpyxl.workbook.workbook, googleapiclient.discovery, functools
import google.oauth2.service_account
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.styles import PatternFill, Border, Alignment, NamedStyle, Side, Font
from openpyxl.utils import get_column_letter
from dataclasses import dataclass, field, InitVar, replace

def warning(s = "Warning"):
    return f"\x1b[38;2;255;255;0m{s}\033[0m"
def message(s = "Message"):
    return f"\x1b[38;2;83;141;213m{s}\033[0m"
def error(s = "Error"):
    return f"\x1b[38;2;255;0;0m{s}\033[0m"

colorama.just_fix_windows_console()
patch_version = sys.argv[1] if len(sys.argv) > 1 else "y_s_"

operators_file_name += ".json"
attachment_overview_file_name += ".json"
xlsx_output_file_name += ".xlsx"

github_link = "https://github.com/hanslhansl/Rainbow-Six-Siege-Weapon-Statistics/"
google_sheets_link = "https://docs.google.com/spreadsheets/d/1QgbGALNZGLlvf6YyPLtywZnvgIHkstCwGl1tvCt875Q"
google_drive_link = "https://docs.google.com/spreadsheets/d/e/2PACX-1vQ1KitQsZksdVP9YPDInxK3xE2gtu1mpUxV5_PNyE8sSm-vFINdbiL8vo9RA2CRSIbIUePLVA1GCTWZ/pubhtml"

tdok_hp_levels = (100, 110, 125, 120, 130, 145)
tdok_with_rook = (False, False, False, True, True, True)
tdok_levels_descriptions = tuple(f"{int(i%3)+1} armor {'+ Rook ' if with_rook else ''}({hp} hp)"
                                 for i, (hp, with_rook) in enumerate(zip(tdok_hp_levels, tdok_with_rook)))
tdok_levels_descriptions_short = tuple(f"{int(i%3)+1}{'R' if with_rook else ''} ({hp})"
                                 for i, (hp, with_rook) in enumerate(zip(tdok_hp_levels, tdok_with_rook)))

# check if the settings are correct
if not os.path.isfile(operators_file_name):
    raise Exception(f"{error()}: '{operators_file_name}' is not a valid file path.")
if not os.path.isdir(weapon_data_dir):
    raise Exception(f"{error()}: '{weapon_data_dir}' is not a valid directory.")
if not 0 <= first_distance:
    raise Exception(f"{error()}: 'first_distance' must be >=0 but is {first_distance}.")
if not first_distance <= last_distance:
    raise Exception(f"{error()}: 'last_distance' must be >='first_distance'={first_distance} but is {last_distance}.")

def deserialize_json(file_name : str):
    with open(file_name, "r", encoding='utf-8') as file:
        try:
            content = json.load(file)
        except json.JSONDecodeError:
            raise Exception(f"{error()}: The json deserialization of file '{file_name}' failed.")
    return content

class RGBA:
    def __init__(self, r : int, g : int, b : int, a : float):
        self.r = r
        self.g = g
        self.b = b
        self.a = a

    @classmethod
    def from_rgb_hex(cls, hex_str : str):
        hex_str = hex_str.lstrip('#')
        if len(hex_str) != 6:
            raise ValueError("Hex string must be 6 characters long.")
        r = int(hex_str[0:2], 16)
        g = int(hex_str[2:4], 16)
        b = int(hex_str[4:6], 16)
        a = 1.0

        return cls(r, g, b, a)
        
    def to_rgb_hex(self, with_hastag = True):
        r = int(self.r + (255 - self.r) * (1 - self.a))
        g = int(self.g + (255 - self.g) * (1 - self.a))
        b = int(self.b + (255 - self.b) * (1 - self.a))
        return f"{'#' if with_hastag else ''}{r:02X}{g:02X}{b:02X}"

    def to_css(self):
        return f"#{self.r:02X}{self.g:02X}{self.b:02X}{int(self.a*0xFF):02X}"
        return f"rgba({self.r}, {self.g}, {self.b}, {self.a})"

    def with_alpha(self, a : float):
        return RGBA(self.r, self.g, self.b, a)

    def to_ex_fill(self):
        rgb = self.to_rgb_hex(False)
        return PatternFill(start_color=rgb, end_color=rgb, fill_type="solid")

    def to_border_color(self, factor=0.8):
        # RGB to HLS
        h, l, s = colorsys.rgb_to_hls(self.r/255, self.g/255, self.b/255)
        l = max(0, min(1, l * factor))  # Helligkeit anpassen
        r_new, g_new, b_new = colorsys.hls_to_rgb(h, l, s)
        
        return RGBA(int(r_new*255), int(g_new*255), int(b_new*255), 1.)

def normalize(min_val, max_val, value):
    # Clamp value within range
    value = max(min_val, min(max_val, value))
    
    # Normalize to [0, 1]
    top = value - min_val
    bottom = max_val - min_val
    if top == bottom:
        return 1.
    return top / bottom
def get_non_stagnant_intervals(data) -> tuple[tuple[int, int],...]:
    intervals = []
    start = None
    for i in range(len(data) - 1):
        if data[i] > data[i + 1]:
            if start is None:
                start = i
        else:
            if start is not None:
                intervals.append((start, i))
                start = None
    if start is not None:
        intervals.append((start, len(data) - 1))
    return tuple(intervals)
def is_index_in_intervals(index : int, intervals : tuple[tuple[int, int],...]):
    for start, end in intervals:
        if start <= index <= end:
            return start, end
    return None

def safe_division(a : float, b : float):
    if a == b:
        return 1.
    return a / b

border_style = "thin"
center_alignment = Alignment("center", wrapText=True)
left_alignment = Alignment("left", wrapText=True)

attachment_overview_json = deserialize_json(attachment_overview_file_name)

class Weapons:
    def __init__(self):
        attachment_categories = deserialize_json(attachment_overview_file_name)
        Weapon.extended_barrel_damage_multiplier = 1.0 + attachment_categories["Barrels"]["Extended barrel"]["damage bonus"]
        Weapon.laser_ads_speed_multiplier = 1.0 + attachment_categories["Under Barrel"]["Laser"]["ads speed bonus"]

        # get operators
        json_content = deserialize_json(operators_file_name)
        if type(json_content) != dict:
            raise Exception(f"{error()}: File '{operators_file_name}' doesn't deserialize to a dict of operators and weapons lists.")
        self.operators = [Operator(js, op_name) for (op_name, js) in json_content.items()]

        # get weapons
        weapons : list[Weapon] = []
        for file_name in os.listdir(weapon_data_dir):
            file_path = os.path.join(weapon_data_dir, file_name)

            name, extension = os.path.splitext(file_name);		
            if not extension == ".json":
                continue
            if name.startswith("_"):
                print(f"{message('Message: Excluding')} weapon '{message(file_name)}' because of _.")
                continue
        
            w = Weapon(file_path, self.operators)
            weapons.append(w)
        
        # verify operator weapons
        for op in self.operators:
            for weapon_name in op._weapons:
                if weapon_name not in op.weapons:
                    print(f"{warning('Warning:')} Weapon '{warning(weapon_name)}' found on operator '{op.name}' is {warning('not an actual weapon')}.")
            del op._weapons

        # add eb weapons
        weapons += (w.extended_barrel_weapon for w in weapons if w.extended_barrel_weapon)
    
        weapons_sorted = sorted(weapons, key=lambda w: (Weapon.classes.index(w.class_), w.name))
        self.weapons = {w.name : w for w in weapons_sorted}
        self.base_weapons = {n : w for n, w in self.weapons.items() if not w.is_extended_barrel}

        self._damages = pd.DataFrame({name : w.damages for name, w in self.weapons.items()}).transpose()
        self._damages.index.rename("weapons", inplace=True)
        for name, w in self.weapons.items():
            del w.damages

        return

    def filter(self, df : pd.DataFrame, filter_func : typing.Callable[["Weapon"], bool]) -> pd.DataFrame:
        return df[df.apply(lambda row: filter_func(self.weapons[row.name]), axis=1)]
    def apply(self, df : pd.DataFrame | pd.io.formats.style.Styler, callback : typing.Callable[["Weapon", typing.Any, int, typing.Any], typing.Any]):
        def cb(x):
            name, pi = (x.name, None) if df.index.nlevels == 1 else x.name
            if pi == None:
                return [None] * len(x)
            w = self.weapons[name]
            return [callback(w, pi, i, v) for i, v in x.items()]
        
        return df.apply(cb, axis=1)
    def apply_style(self, df : pd.DataFrame, callback : typing.Callable[["Weapon", typing.Any, int, typing.Any], str]):
        return self.apply(df.style, callback)
    def apply_background_color(self, stat : "Stat", callback : typing.Callable[["Weapon", typing.Any, int, typing.Any], RGBA]):
        convert_color = RGBA.to_rgb_hex if __name__ == "__main__" else RGBA.to_css
        def format_value(val):
            if pd.isna(val):
                return ""
            elif isinstance(val, float):
                return f"{val:.2f}".rstrip('0').rstrip('.')
            return val
        def format_index(val):
            if isinstance(val, str):
                return val
            return stat.additional_parameters_short_description[val%(len(stat.additional_parameters)+1) - 1]

        return (self.apply_style(stat.data, lambda *args: f"background-color: {convert_color(callback(*args))}")
                .format(format_value)
                #.relabel_index([format_index(x) for x in stat.target.index], axis=0) # not working
                )

    # stats helper
    def is_in_damage_drop_off(self, w : "Weapon", index : int):
        return is_index_in_intervals(index, w.damage_drop_off_intervals) is None
    
    def extended_barrel_difference(self, stat : "Stat"):
        """this has to loop over all params, maybe change vectorize_and_interleave to differentiate between 1 param and >1 params"""
        def impl(p):
            df = stat.source[stat.additional_parameters.index(p)]
            has_or_is_eb = self.filter(df, lambda w: w.is_extended_barrel or w.extended_barrel_weapon != None)
            prim = self.filter(has_or_is_eb, lambda w: w.extended_barrel_weapon != None)
            sec = self.filter(has_or_is_eb, lambda w: w.is_extended_barrel)

            pd.options.mode.chained_assignment, old = None, pd.options.mode.chained_assignment
            has_or_is_eb.loc[prim.index] -= prim.values
            has_or_is_eb.loc[sec.index] -= prim.values
            pd.options.mode.chained_assignment = old

            target = (prim - sec.values).abs()
            source = has_or_is_eb

            return target, source
        
        target, source, _ = self.vectorize_and_interleave(impl, stat.additional_parameters)
        return replace(stat, target=target, source=source)
        

    # illustrations helper
    def vectorize_and_interleave(self, func : typing.Callable[[typing.Any], pd.DataFrame], params : tuple):
        sources : list[pd.DataFrame] = []
        targets : list[pd.DataFrame] = []
        for p in params:
            x = func(p)
            target, source = (x, x) if isinstance(x, pd.DataFrame) else x
            sources.append(source)
            targets.append(target)
        if len(params) == 1:
            target = targets[0]
        else:
            dfs = [pd.DataFrame(np.nan, index=targets[0].index, columns=targets[0].columns)] + targets
            interleaved_rows = [row for pair in zip(*(df.iterrows() for df in dfs)) for row in pair]
            target = pd.DataFrame([row[1] for row in interleaved_rows])
            target.index = [(i if i%(len(params)+1)!=0 else s) for i, s in enumerate(target.index)]
        return target, tuple(sources), params

    def nest(self, func : typing.Callable[[typing.Any], pd.DataFrame], params : tuple, pname : str):
        names = tuple(range(len(params)))
        res = pd.concat([func(p) for p in params], keys=params, names=[pname])
        res.index = pd.MultiIndex.from_product([self.weapons, names], names=["weapons", pname])
        index = pd.MultiIndex.from_product([self.weapons, ("",) + names], names=["weapons", pname])
        return res.reindex(index), params

    """
        return Stat(
            "",
            "",
            "",
            *self.nest(lambda p: , tdok_hp_levels, hp),
            tdok_levels_descriptions_short,
            tdok_levels_descriptions
            )
    """

    # primary stats
    @functools.cache
    def damage_per_bullet(self):
        return Stat("damage per bullet", "damage per bullet", "damage-per-bullet", self._damages)
    @functools.cache
    def damage_per_shot(self):
        pellets = {name : w.pellets for name, w in self.weapons.items()}
        res = self._damages.mul(pellets, axis=0)
        return Stat("damage per shot", "damage per shot", "damage-per-shot", res)
    @functools.cache
    def dps(self):
        """damage per second"""
        bullets_per_second = {name : w.pellets * w.rps for name, w in self.weapons.items()}
        res = self._damages.mul(bullets_per_second, axis=0)
        return Stat("dps", "damage per second", "damage-per-second---dps", res)
    
    @functools.cache
    def btdok(self):
        return Stat(
            "btdok",
            "bullets to down or kill",
            "bullets-to-down-or-kill---btdok",
            *self.nest(lambda hp: np.ceil(hp / self._damages), tdok_hp_levels, "hp"),
            tdok_levels_descriptions_short,
            tdok_levels_descriptions
            )
    @functools.cache
    def stdok(self):
        pellets = {name : w.pellets for name, w in self.weapons.items()}
        return Stat(
            "stdok",
            "shots to down or kill",
            "shots-to-down-or-kill---stdok",
            *self.nest(lambda hp: np.ceil((hp / self._damages).div(pellets, axis=0)), tdok_hp_levels, "hp"),
            tdok_levels_descriptions_short,
            tdok_levels_descriptions
            )
        return np.ceil((hp / self._damages).div(pellets, axis=0))
    @functools.cache
    def ttdok(self):
        pellets = {name : w.pellets for name, w in self.weapons.items()}
        rpms = {name : w.rpms for name, w in self.weapons.items()}
        return Stat(
            "ttdok",
            "time to down or kill",
            "time-to-down-or-kill---ttdok",
            *self.nest(lambda hp: (np.ceil((hp / self._damages).div(pellets, axis=0)) - 1).div(rpms, axis=0).round(), tdok_hp_levels, "hp"),
            tdok_levels_descriptions_short,
            tdok_levels_descriptions
            )
    
    @functools.cached_property
    def theoretical_btdok(self, hp : int):
        """"""
        return hp / self._damages
    @functools.cached_property
    def theoretical_stdok(self, hp : int):
        """"""
        pellets = {name : w.pellets for name, w in self.weapons.items()}
        return (hp / self._damages).div(pellets, axis=0)
    @functools.cached_property
    def theoretical_ttdok(self, hp : int):
        """"""
        pellets_rpms = {name : w.pellets * w.rpms for name, w in self.weapons.items()}
        return (hp / self._damages).div(pellets_rpms, axis=0)

    @functools.cached_property
    def btk(self, hp : int):
        """bullets to kill"""
        return self.btdok(hp + 20)
    @functools.cached_property
    def stk(self, hp : int):
        """shots to kill"""
        return self.stdok(hp + 20)
    @functools.cached_property
    def ttk(self, hp : int):
        """time to kill"""
        return self.ttdok(hp + 20)
    
    # illustrations helper
    @functools.cache
    def vectorize_and_interleave1(self, func : typing.Callable[["Weapons", typing.Any], pd.DataFrame], params : tuple, filter : tuple[str] = None):
        sources : list[pd.DataFrame] = []
        targets : list[pd.DataFrame] = []
        if filter is not None: filter = list(filter)
        for p, d, sd in params:
            x = func(self, p)
            target, source = (x, x) if isinstance(x, pd.DataFrame) else x
            sources.append(source)
            if filter is None:
                targets.append(target)
            else:
                targets.append(target.loc[filter])
        params_len = len(params)
        if params_len == 1:
            target = targets[0]
        else:
            dfs = [pd.DataFrame(np.nan, index=targets[0].index, columns=targets[0].columns)] + targets
            interleaved_rows = [row for pair in zip(*(df.iterrows() for df in dfs)) for row in pair]
            target = pd.DataFrame([row[1] for row in interleaved_rows])
            target.index = [(i if i%(params_len+1)!=0 else s) for i, s in enumerate(target.index)]
        target.rename_axis
        return target, sources

    # illustrations
    def damage_drop_off_coloring(self, stat : "Stat"):
        """the colored areas represent steady damage, the colorless areas represent decreasing damage"""
        return self.apply_background_color(
            stat,
            lambda w, i, pi: w.color if self.is_in_damage_drop_off(w, i) else w.empty_color,
            )
    
    def stat_to_base_stat_gradient_coloring(self, data : tuple[pd.DataFrame, list[pd.DataFrame]], params : tuple):
        """the color gradient illustrates the {stat} compared to the weapon's base {stat} (i.e. at 0 m)"""
        target, dfs = data
        return self.apply_background_color(
            target,
            lambda w, i, pi: w.color.with_alpha(safe_division(dfs[pi].loc[w.name][i], dfs[pi].loc[w.name][0])),
            params
            )
    def stat_to_stat_max_gradient_coloring(self, data : tuple[pd.DataFrame, list[pd.DataFrame]], params : tuple):
        """the color gradient illustrates the {stat} compared to the weapon's maximum {stat}"""
        target, dfs = data
        return self.apply_background_color(
            target,
            lambda w, i, pi: w.color.with_alpha(safe_division(dfs[pi].loc[w.name][i], dfs[pi].loc[w.name].max())),
            params
            )

    def stat_to_class_stat_gradient_coloring(self, data : tuple[pd.DataFrame, list[pd.DataFrame]], params : tuple):
        """the color gradient illustrates the {stat} compared to the weapon class' maximum {stat} at the same distance"""
        target, dfs = data
        class_max = [{class_ : group_df.max(axis=0) for class_, group_df in df.groupby(lambda name: self.weapons[name].class_)} for pi, df in enumerate(dfs)]
        return self.apply_background_color(
            target,
            lambda w, i, pi: w.color.with_alpha(safe_division(dfs[pi].loc[w.name][i], class_max[pi][w.class_][i])),
            params
            )

    def stat_to_class_base_stat_gradient_coloring(self, data : tuple[pd.DataFrame, list[pd.DataFrame]], params : tuple):
        """the color gradient illustrates the {stat} compared to the weapon's class' maximum base {stat} (i.e. at 0 m)"""
        target, dfs = data
        class_base_max = [{class_ : group_df.iloc[:, 0].max() for class_, group_df in df.groupby(lambda name: self.weapons[name].class_)} for pi, df in enumerate(dfs)]
        return self.apply_background_color(
            target,
            lambda w, i, pi: w.color.with_alpha(safe_division(dfs[pi].loc[w.name][i], class_base_max[pi][w.class_])),
            params
            )
    def stat_to_class_stat_max_gradient_coloring(self, data : tuple[pd.DataFrame, list[pd.DataFrame]], params : tuple):
        """the color gradient illustrates the {stat} compared to the weapon's class' maximum {stat}"""
        target, dfs = data
        class_max = [{class_ : group_df.to_numpy().max() for class_, group_df in df.groupby(lambda name: self.weapons[name].class_)} for pi, df in enumerate(dfs)]
        return self.apply_background_color(
            target,
            lambda w, i, pi: w.color.with_alpha(safe_division(dfs[pi].loc[w.name][i], class_max[pi][w.class_])),
            params
            )
    
    def extended_barrel_effect_coloring(self, data : tuple[pd.DataFrame, list[pd.DataFrame]], params : tuple):
        """the colored areas show where the extended barrel attachment actually affects the {stat}"""
        target, dfs = data
        pred = (lambda w, i, pi: (w.is_extended_barrel and dfs[pi].loc[w.name][i] != dfs[pi].loc[w.base_name][i])
                or (w.extended_barrel_weapon and dfs[pi].loc[w.name][i] != dfs[pi].loc[w.extended_barrel_weapon.name][i]))
        return self.apply_background_color(
            target,
            lambda w, i, pi: w.color if pred(w, i, pi) else w.empty_color,
            params
            )


@dataclasses_json.dataclass_json
@dataclass
class _Weapon:
    name : str
    class_ : str = field(metadata=dataclasses_json.config(field_name="class"))
    rpm : int
    ads_time : float
    pellets : int
    _capacity : tuple[int, int] = field(metadata=dataclasses_json.config(field_name="capacity")) # magazine, chamber
    extra_ammo : int
    _damages : dict[int, int] = field(metadata=dataclasses_json.config(field_name="damages"))
    has_laser : bool
    has_grip : bool
    #reload_times : tuple[float, float]
    _extended_barrel : dict[str, int] | bool = field(metadata=dataclasses_json.config(field_name="extended_barrel"))

class Weapon(_Weapon):
    colors = {class_: RGBA.from_rgb_hex(color) for class_, color in weapon_colors.items()}
    classes = tuple(colors)
    distances = list(range(first_distance, last_distance+1))

    # excel stuff
    ex_borders = {class_ : Border(
                left=  Side(border_style=border_style, color=color.to_border_color().to_rgb_hex(False)),
                right= Side(border_style=border_style, color=color.to_border_color().to_rgb_hex(False)),
                top=   Side(border_style=border_style, color=color.to_border_color().to_rgb_hex(False)),
                bottom=Side(border_style=border_style, color=color.to_border_color().to_rgb_hex(False))
                ) for class_, color in colors.items()}
    ex_borders = {class_ : Border(*(Side(border_style=border_style, color=color.to_border_color().to_rgb_hex(False)) for i in range(4)))
                  for class_, color in colors.items()}

    extended_barrel_damage_multiplier = 0.0
    laser_ads_speed_multiplier = 0.0
    angled_grip_reload_speed_multiplier = 0.0

    empty_color = RGBA(0,0,0,0)

    def __init__(self, file_path, operators : list["Operator"]):
        json_content = deserialize_json(file_path)

        try:
            with warnings.catch_warnings():
                warnings.simplefilter("ignore", category=UserWarning)
                _w = _Weapon.schema().load(json_content)
        except marshmallow.exceptions.ValidationError as e:
            raise Exception(f"{error()}: File '{file_path}' could not be deserialized: {str(e)}.")
        super().__init__(**vars(_w))

        # get operators
        self.operators : list[Operator] = []
        for op in operators:
            if self.name in op._weapons:
                self.operators.append(op)
                op.weapons.append(self)
                op._weapons.remove(self.name)

        # operators rich text
        self.ex_operators_rich_text = CellRichText([elem for op in self.operators for elem in (op.rich_text_name, ", ")][:-1])

        # get weapon names
        
        
        # verify weapon class
        if self.class_ not in self.classes:
            raise Exception(f"{error()}: Weapon '{self.name}' has an invalid weapon class '{json_content["class"]}'.")
        
        # derived fields
        self.base_name = self.name
        self.display_name = self.name
        self.color = self.colors[self.class_]
        self.name_color = self.color
        self.rps = self.rpm / 60.
        self.rpms = self.rpm / 60000.
        self.ads_time_with_laser = self.ads_time / self.laser_ads_speed_multiplier if self.has_laser else None
        self.capacity = str(self._capacity[0]) + "+" + str(self._capacity[1])
        #self.reload_times_with_angled_grip = tuple(x / self.angled_grip_reload_speed_multiplier for x in self.reload_times) if self.has_grip else None

         # excel fields
        self.ex_border = self.ex_borders[self.class_]

        # verify weapon damages
        setattr(self, "damages", self.validate_damages(self._damages))
        self.damage_drop_off_intervals = self.get_damage_drop_off_intervals()

        # get extended barrel weapon, needs to be last bc of copy()
        self.extended_barrel_parent : Weapon | None = None
        self.is_extended_barrel = False
        self.extended_barrel_weapon = None
        if self._extended_barrel != False:
            self.extended_barrel_weapon = copy.copy(self)
            self.extended_barrel_weapon.name = self.name + " + eb"
            self.extended_barrel_weapon.display_name = "+ extended barrel"
            self.extended_barrel_weapon.name_color = self.empty_color
            setattr(
                self.extended_barrel_weapon,
                "damages",
                self.extended_barrel_weapon.validate_damages({int(k) : v for k, v in self._extended_barrel.items()}))
            self.extended_barrel_weapon.damage_drop_off_intervals = self.extended_barrel_weapon.get_damage_drop_off_intervals()

            self.extended_barrel_weapon.extended_barrel_parent = self
            self.extended_barrel_weapon.is_extended_barrel = True

        return

    def __repr__(self):
        return f"<Weapon: {self.name}>"

    def validate_damages(self, distance_damage_dict : dict[int, int]):
        # insert missing distances with damage = 0
        for distance in Weapon.distances:
            if distance not in distance_damage_dict:
                distance_damage_dict[distance] = 0
                
        # sort distance_damage_dict in ascending order by distance
        distance_damage_dict = dict(sorted(distance_damage_dict.items()))

        distances = list(distance_damage_dict.keys())
        damages = list(distance_damage_dict.values())
        
        #if Weapon.distances != distances:
        if Weapon.distances != distances:
            raise Exception(f"{error()}: Weapon '{self.name}' has incorrect distance values.")

        # make sure the last damage value is given. otherwise the extrapolation will be wrong
        if damages[-1] == 0:
            raise Exception(f"{error()}: Weapon '{self.name}' is missing a damage value at {distances[-1]}m.")

        # make sure damages only stagnate or decrease and zeros are surrounded by identical non-zero damages
        # interpolate gaps. damages will be continuous in [5;40]
        previous_real_damage = 0
        previous_was_interpolated = False
        for i in range(len(Weapon.distances)):
            if damages[i] == 0:	# this damage value needs to be interpolated
                damages[i] = previous_real_damage
                previous_was_interpolated = True
                
            else:	# this damage value is given
                if damages[i] > previous_real_damage and previous_real_damage != 0:
                    raise Exception(f"{error()}: Weapon '{self.name}' has a damage increase from '{previous_real_damage}' to '{damages[i]}' at {Weapon.distances[i]}m.")
                if previous_real_damage != 0 and previous_was_interpolated == True and damages[i] != previous_real_damage:
                    raise Exception(f"{error()}: Tried to interpolate between two unequal damage values '{previous_real_damage}' and '{damages[i]}' at {Weapon.distances[i]}m for weapon '{self.name}'.")
                
                previous_real_damage = damages[i]
                previous_was_interpolated = False

        # get index to first non-zero damage
        first_nonzero_index = next((i for i, damage in enumerate(damages) if damage != 0), -1)
        
        if first_nonzero_index > 5:
            raise Exception(f"{error()}: First non-zero damage value for weapon '{self.name}' is at {Weapon.distances[first_nonzero_index]}m. Should be at 5m or less.")

        # extrapolate first 5 meters. damages will be continuous in [0;40]
        if first_nonzero_index == 0:
            pass	# no extrapolation needed
        elif first_nonzero_index == -1:
            raise Exception(f"{error()}: Weapon '{self.name}' has no damage values at all.")
        else:
            if self.class_ == "SG" or self.name == "Glaive-12":	# special treatment for shotgunsand glaive-12
                if first_nonzero_index <= 5:
                    for i in range(first_nonzero_index):
                        damages[i] = damages[first_nonzero_index]
                else:
                    raise Exception(f"{error()}: Can't extrapolate first {first_nonzero_index} meters for shotgun '{self.name}'.")
            else:
                if damages[first_nonzero_index] == damages[first_nonzero_index+1] == damages[first_nonzero_index+2]:
                    for i in range(first_nonzero_index):
                        damages[i] = damages[first_nonzero_index]
                else:
                    raise Exception(f"{error()}: Can't extrapolate first {first_nonzero_index} meters for weapon '{self.name}'.")

        # return the damage stats
        return tuple(damages)

    def get_damage_drop_off_intervals(self):
        intervals = get_non_stagnant_intervals(self.damages)
        if self.class_ == "SG":
            if len(intervals) != 2:
                raise Exception(f"{error()}: A {self.class_} should have exactly 2 damage dropoff intervals but weapon '{self.name}' has {len(intervals)}.")
            return intervals
        return (intervals[0][0], intervals[-1][-1]),
    

@dataclasses_json.dataclass_json
@dataclass
class _Operator:
    _side : str = field(metadata=dataclasses_json.config(field_name="side"))
    _weapons : list[str] = field(metadata=dataclasses_json.config(field_name="weapons"))

class Operator(_Operator):
    attacker_color = RGBA.from_rgb_hex("#198FEB")
    defender_color = RGBA.from_rgb_hex("#FB3636")

    def __init__(self, json_content, name : str):
        self.name = name

        try:
            _w = _Operator.schema().load(json_content)
        except marshmallow.exceptions.ValidationError as e:
            raise Exception(f"{error()}: Operator '{self.name}' could not be deserialized: {str(e)}.")
        super().__init__(**vars(_w))

        if self._side not in ("A", "D"):
            raise Exception(f"{error()}: Operator '{self.name}' has an invalid side value '{self._side}'.")
        self.side = bool(self._side == "D")	# False: attack, True: defense
        
        self.weapons : list[Weapon] = []
        
        self.rich_text_name = TextBlock(InlineFont(color=Operator.defender_color.to_rgb_hex(False) if self.side else Operator.attacker_color.to_rgb_hex(False)), self.name)

        return


@dataclass
class Stat:
    short_name : str
    name : str
    link : InitVar[str]
    data : pd.DataFrame
    additional_parameters : tuple[typing.Any,...] = None,
    additional_parameters_short_description : tuple[str,...] = "",
    additional_parameters_description : tuple[str,...] = "",

    def __post_init__(self, link : str):
        self.link = f"https://github.com/hanslhansl/Rainbow-Six-Siege-Weapon-Statistics/#{link}"

    @property
    def display_name(self):
        if self.short_name != self.name:
            return self.short_name + " - " + self.name
        return self.name

@dataclass
class Stat2:
    _link : str
    stat_method : typing.Callable[[Weapons, typing.Any], pd.DataFrame]
    is_tdok : bool

    def __post_init__(self):
        self.link = f"https://github.com/hanslhansl/Rainbow-Six-Siege-Weapon-Statistics/#{self._link}"

        self.short_name = self.stat_method.__name__.replace("_", " ")
        self.name = self.short_name if self.stat_method.__doc__ is None else self.stat_method.__doc__

        if self.is_tdok:
            self.additional_parameter_name = "hp"
            self.additional_parameters = tuple(zip(tdok_hp_levels, tdok_levels_descriptions, tdok_levels_descriptions_short))
        else:
            self.additional_parameter_name = ""
            self.additional_parameters = (None, "", ""),

    @property
    def display_name(self):
        if self.short_name != self.name:
            return self.short_name + " - " + self.name
        return self.name


stats = (
    Weapons.damage_per_bullet,
    Weapons.damage_per_shot,
    Weapons.dps,
    
    Weapons.btdok,
    Weapons.stdok,
    Weapons.ttdok,
)
"""Stat("damage-per-bullet", Weapons.damage_per_bullet, False),
    Stat("damage-per-shot", Weapons.damage_per_shot, False),
    Stat("damage-per-second---dps", Weapons.dps, False),

    Stat("bullets-to-down-or-kill---btdok", Weapons.btdok, True),
    Stat("shots-to-down-or-kill---stdok", Weapons.stdok, True),
    Stat("time-to-down-or-kill---ttdok", Weapons.ttdok, True),

    Stat("bullets-to-down-or-kill---btdok", Weapons.btk, True),
    Stat("shots-to-down-or-kill---stdok", Weapons.stk, True),
    Stat("time-to-down-or-kill---ttdok", Weapons.ttk, True),

    Stat("bullets-to-down-or-kill---btdok", Weapons.theoretical_btdok, True),
    Stat("shots-to-down-or-kill---stdok", Weapons.theoretical_stdok, True),
    Stat("time-to-down-or-kill---ttdok", Weapons.theoretical_ttdok, True),"""
stat_illustrations = (
    Weapons.damage_drop_off_coloring,
    #Weapons.stat_to_base_stat_gradient_coloring,
    Weapons.stat_to_stat_max_gradient_coloring,
    Weapons.stat_to_class_stat_gradient_coloring,
    #Weapons.stat_to_class_base_stat_gradient_coloring,
    Weapons.stat_to_class_stat_max_gradient_coloring,
    Weapons.extended_barrel_effect_coloring,
    )

def add_worksheet_header(worksheet : typing.Any, stat : Stat | str, description : str, row : int, cols_inbetween : int):
    
    worksheet.column_dimensions[get_column_letter(1)].width = 22
    
    def add_header_entry(row, start_column, end_column, value, font = None):
        worksheet.merge_cells(start_row=row, end_row=row, start_column=start_column, end_column=end_column)
        c = worksheet.cell(row=row, column=start_column)
        c.value = value
        if font is not None:
            c.font = font

    add_header_entry(row, 2, 1 + cols_inbetween,
     f"created by hanslhansl, updated for {patch_version}", Font(bold=True))
    row += 1

    add_header_entry(row, 2, 6, f'=HYPERLINK("{github_link}", "detailed explanation")', Font(color = "FF0000FF"))

    add_header_entry(row, 8, 14, f'=HYPERLINK("{google_sheets_link}", "spreadsheet on google sheets")', Font(color = "FF0000FF"))

    add_header_entry(row, 16, 1 + cols_inbetween,
                  f'=HYPERLINK("{google_drive_link}", "spreadsheet on google drive")', Font(color = "FF0000FF"))
    row += 2

    if isinstance(stat, str):
        add_header_entry(row, 2, 1 + cols_inbetween, stat, Font(color = "FF0000FF", bold=True))
    else:
        add_header_entry(row, 2, 1 + cols_inbetween,
                         f'=HYPERLINK("{github_link}#{stat.link}", "{stat.display_name}")', Font(color = "FF0000FF", bold=True))
    row += 1

    add_header_entry(row, 2, 1 + cols_inbetween, description)
    row += 1

    return row

def add_secondary_weapon_stats_header(worksheet : typing.Any, row : int, col : int):
    
    empty = (None, 3)
    values_widths = (
        empty,
        ("class", 10),
        empty,
        ("rpm", 6),
        ("capacity", 10),
        ("ammo", 8),
        ("pellets", 8),
        empty,
        ("ads", 6),
        ("+ laser", 9),
        empty,
        ("full reload", 11),
        ("tactical", 8),
        ("+ angled grip", 7),
        (None, 7),
        empty,
        ("operators", 50)
    )

    for value, width in values_widths:
        if value != None:
            c = worksheet.cell(row=row, column=col)
            c.value = value
            c.alignment = center_alignment
        worksheet.column_dimensions[get_column_letter(col)].width = width

        col += 1

    worksheet.cell(row=row, column=col-1).alignment = left_alignment

    worksheet.merge_cells(start_row=row, end_row=row, start_column=col-4, end_column=col-4+1)

    return row, col

def add_secondary_weapon_stats(worksheet : typing.Any, weapon : Weapon, row : int, col : int):
    values = [weapon.class_, weapon.rpm, weapon.capacity, weapon.extra_ammo, weapon.pellets if weapon.pellets != 1 else None,
           weapon.ads_time, weapon.ads_time_with_laser,
           #Weapon.getReload, Weapon.getReloadTimesWithAngledGrip
          ]
    skips = [2, 1, 1, 1, 2, 1, 7]
    col += 1

    for value, skip in zip(values, skips):
        c = worksheet.cell(row=row, column=col)
        
        if value != None:
            if type(value) == float:
                value = round(value, 3)
            c.value = value
            c.fill = weapon.color.to_ex_fill()
            c.border = weapon.ex_border
            c.alignment = center_alignment
            weapon.ex_border
        col += skip

    c1 = worksheet.cell(row=row, column=col)
    c1.value = weapon.ex_operators_rich_text

    return

def add_extended_barrel_overview(worksheet : typing.Any, ws : Weapons, row : int, col : int, with_secondary_weapon_stats : bool):
    original_row = row

    c = worksheet.cell(row=row, column=col)
    c.value = "weapon"
    worksheet.column_dimensions[get_column_letter(col)].width = 22

    col += 1
    selected_stats = [stats[x] for x in (4, 5)]
    # loop over stats, stdok and ttdok
    for i, stat in enumerate(selected_stats):
        stat = stat(ws)
        row = original_row

        # stat name
        worksheet.merge_cells(start_row=row, end_row=row, start_column=col, end_column=col+len(stat.additional_parameters)-1)
        c = worksheet.cell(row=row, column=col)
        c.value = stat.short_name
        c.alignment = center_alignment
        row += 1

        # loop over health ratings
        for j, (param, sd) in enumerate(zip(stat.additional_parameters, stat.additional_parameters_short_description)):
            row = original_row

            # secondary weapon stats header
            secondary_column_offset = len(selected_stats)*(len(stat.additional_parameters)+1) - 1
            if i == 0 and j == 0 and with_secondary_weapon_stats:
                add_secondary_weapon_stats_header(worksheet, row, col + secondary_column_offset)

            row += 1

            # health rating
            c = worksheet.cell(row=row, column=col)
            c.value = sd
            c.alignment = center_alignment
            worksheet.column_dimensions[get_column_letter(col)].width = 9

            # loop over weapons
            target = ws.extended_barrel_difference(stat).target
            for name, row_data in target.iterrows():
                w = ws.base_weapons[name]

                # weapon name cell
                if i == 0 and j == 0:
                    c = worksheet.cell(row=row, column=col-1)
                    c.value = name
                    c.fill = w.name_color.to_ex_fill()
                    c.border = w.ex_border

                    # secondary weapon stats
                    if with_secondary_weapon_stats:
                        add_secondary_weapon_stats(worksheet, w, row, col+secondary_column_offset)

                # data cell
                c = worksheet.cell(row=row, column=col)
                c.value = row_data[0]
                c.fill = (w.color if row_data[0] != 0 else w.empty_color).to_ex_fill()
                c.alignment = center_alignment
                c.border = w.ex_border

                row += 1
            col += 1
            worksheet.column_dimensions[get_column_letter(col)].width = 3
        col += 1

    return row

def add_attachment_overview(workbook : typing.Any, ws : Weapons):
    json_content = deserialize_json(attachment_overview_file_name)
    
    if not isinstance(json_content, dict):
        raise Exception(f"{error()}: File '{attachment_overview_file_name}' doesn't deserialize to a dictionary.")
    attachment_categories : dict[str, typing.Any] = json_content

    worksheet = workbook.create_sheet("attachments")
    row = add_worksheet_header(worksheet, "attachment overview", "a short overview over all available attachments.", 1, 19)
    #worksheet.freeze_panes = worksheet.cell(row=row, column=2)

    for attachment_category, attachment_dict in attachment_categories.items():
        if not isinstance(attachment_dict, dict):
            raise Exception(f"{error()}: An attachment category in file '{attachment_overview_file_name}' doesn't deserialize to a dictionary but to '{type(attachment_dict)}'.")
        attachment_dict : dict[str, typing.Any]

        c = worksheet.cell(row=row, column=1)
        c.value = attachment_category
        c.font = Font(bold=True)
        
        for attachment_name, attachment in attachment_dict.items():
            if not isinstance(attachment, dict):
                raise Exception(f"{error()}: An attachment in file '{attachment_overview_file_name}' doesn't deserialize to a dictionary but to '{type(attachment)}'.")

            worksheet.merge_cells(start_row=row, end_row=row, start_column=2, end_column=1 + 19)
            c = worksheet.cell(row=row, column=2)
            c.value = attachment_name
            c.font = Font(bold=True)
            row += 1
            
            if "description" in attachment:
                description = attachment.pop("description").format(*attachment.values())
                worksheet.merge_cells(start_row=row, end_row=row, start_column=2, end_column=1 + 19)
                c = worksheet.cell(row=row, column=2)
                c.value = description
                row += 1
                
            if "damage bonus" in attachment:
                row += 1
                row = add_extended_barrel_overview(worksheet, ws, row, 2, True)
                
            row += 1
            
        row += 1

def modify_stats_worksheet(worksheet : typing.Any, ws : Weapons, stat : Stat, illustration):
    row = 1

    c = worksheet.cell(row=row, column=1)
    c.value = "weapon"
    worksheet.column_dimensions[get_column_letter(1)].width = 22

    col = 2
    for d in Weapon.distances:
        c = worksheet.cell(row=row, column=col)
        c.value = d
        c.alignment = center_alignment
        worksheet.column_dimensions[get_column_letter(col)].width = 4.8
        col += 1

    row, _ = add_secondary_weapon_stats_header(worksheet, row, col)
    border = Border(bottom=Side(style="thick"))
    for c in range(1, col):
        worksheet.cell(row=row, column=c).border = border
    worksheet.freeze_panes = worksheet.cell(row=row+1, column=2)
    row += 2

    row = add_worksheet_header(worksheet, stat, illustration.__doc__.format(stat=stat.short_name), row, len(Weapon.distances))
    row += 1

    is_1d_stat = len(stat.additional_parameters) == 1
    for weapon in ws.weapons.values():
        # weapon name cell
        c = worksheet.cell(row=row, column=1)
        c.value = weapon.display_name
        c.style = "Normal"
        c.fill = weapon.name_color.to_ex_fill()
        c.border = weapon.ex_border

        # secondary weapn stats
        if not weapon.is_extended_barrel:
            add_secondary_weapon_stats(worksheet, weapon, row, col)

        if not is_1d_stat: row += 1
        for param_desc in stat.additional_parameters_description:
            if not is_1d_stat:
                # parameter description cell
                c = worksheet.cell(row=row, column=1)
                c.value = param_desc
                c.style = "Normal"
                c.border = weapon.ex_border

            # stat cells
            for i in range(len(Weapon.distances)):
                c = worksheet.cell(row=row, column=i+2)
                c.alignment = center_alignment
                c.border = weapon.ex_border

            row += 1

    return


def save_to_xlsx_file(ws : Weapons):
    """ https://openpyxl.readthedocs.io/en/stable/ """

    stat_indices = (0, 1, 2, 4, 5)
    illustration_indices = (0, 1, 2, 4, 3)
    illustration_indices = (0, 0, 0, 0, 0)

    excel_buffer = io.BytesIO()
    with pd.ExcelWriter(excel_buffer) as writer:
        print("pandas to excel engine:", writer.engine)

        for i_stat, i_illustration in zip(stat_indices, illustration_indices):
            stat = stats[i_stat](ws)
            illustration = stat_illustrations[i_illustration]

            data = stat.data #ws.vectorize_and_interleave(stat.stat_method, stat.additional_parameters)
            styler = illustration(ws, stat)
            styler.to_excel(writer, sheet_name=stat.short_name, header=False, startrow=8)

    workbook = openpyxl.load_workbook(excel_buffer)

    for i_stat, i_illustration in zip(stat_indices, illustration_indices):
        stat = stats[i_stat](ws)
        illustration = stat_illustrations[i_illustration]
        worksheet = workbook[stat.short_name]

        #modify_stats_worksheet(worksheet, ws, stat, illustration)

    #add_attachment_overview(workbook, ws)

    # save to file
    workbook.save(xlsx_output_file_name)
    
    return xlsx_output_file_name

if __name__ == "__main__":

    # get all weapons from the files
    ws = Weapons()

    # verify
    # group weapons by class and by base damage
    """weapons_sorted = sorted((w for w in ws.weapons.values() if not w.is_extended_barrel), key=lambda w: (w.class_, ws.damage_per_bullet().data[0][w.name]))
    grouped = [list(group) for key, group in itertools.groupby(weapons_sorted, key=lambda w: (w.class_, ws.damage_per_bullet().data[0][w.name]))]
    # find all weapons with the same base damage but different damage drop-off
    failed = False
    for group in grouped:
        if len(group) > 1:
            for i, distance in enumerate(Weapon.distances):
                if len(set(ws.damage_per_bullet().data[i][weapon.name] for weapon in group)) > 1:
                    print(f"{warning()}: These {group[0].class_}s have the {warning('same base damage')} ({ws.damage_per_bullet().data[0][group[0].name]}) but {warning('different damages')} at {distance}m:")
                    for weapon in group:
                        print(f"{weapon.name}: {ws.damage_per_bullet().data[i][weapon.name]}")
                    failed = True
    if failed: raise Exception(f"{error()}: See above warnings.")"""


    # save to excel file
    excel_file_name = save_to_xlsx_file(ws)

    if "GOOGLE_SERVICE_ACCOUNT_CREDENTIALS" in os.environ:
        # we are in the github action
        credentials_json = os.environ["GOOGLE_SERVICE_ACCOUNT_CREDENTIALS"]
        info = json.loads(credentials_json)

        # create service account credentials
        creds = google.oauth2.service_account.Credentials.from_service_account_info(
            info,
            scopes=["https://www.googleapis.com/auth/drive"]
        )

        drive_service = googleapiclient.discovery.build("drive", "v3", credentials=creds)

        # Google Sheet ID
        sheet_id = "1QgbGALNZGLlvf6YyPLtywZnvgIHkstCwGl1tvCt875Q"

        # upload excel file replace sheet
        file_metadata = {"mimeType": "application/vnd.google-apps.spreadsheet"}
        media = googleapiclient.http.MediaFileUpload(
            excel_file_name,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            resumable=True
        )

        updated_file = drive_service.files().update(
            fileId=sheet_id,
            media_body=media,
            body=file_metadata,
            fields="id"
        ).execute()

    else:
        # we are local
        os.system("start " + excel_file_name)

